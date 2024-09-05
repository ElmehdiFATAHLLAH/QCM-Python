from openpyxl import load_workbook,Workbook
from openpyxl.styles import Alignment,Border, Side,PatternFill,Font
import tkinter as tk
from datetime import datetime
from tkinter import font
from threading import Thread,Event
from PIL import ImageTk,Image
import random 
import fonctions
import time
import os
import string

center = Alignment(horizontal="center", vertical="center",wrapText=True)
left = Alignment(horizontal="left", vertical="center",wrapText=True)
thick= Border(
    left=Side(border_style="medium", color="000000"),
    right=Side(border_style="medium", color="000000"),
    top=Side(border_style="medium", color="000000"),
    bottom=Side(border_style="medium", color="000000") )
green=Font(name='Calibri', size=12, bold=True,color="008000")
red=Font(name='Calibri', size=12, bold=True,color="FF0000")
bold=Font(bold=True)
yellow= PatternFill(start_color="FFFF00",end_color="FFFF00", fill_type="solid")
bleu= PatternFill(start_color="ADD8E6",end_color="ADD8E6", fill_type="solid")
white= PatternFill(start_color="FFFFFF",end_color="FFFFFF", fill_type="solid")

def minuteur(event,s,num):
    for i in range(s,0,-1):
        if i/10<1:
            label_compteur.config(text=f"Minuteur : 00:00:0{i}")
        else:
            label_compteur.config(text=f"Minuteur : 00:00:{i}")
        time.sleep(1)
        if event.is_set():
            return        
    if num==1:
        suivant()
    else:
        suivant_direct()
    
def afficher_question(dict_question, liste_question, index):
    global event, flag,helpCompteur
    flag=0
    question=liste_question[index]
    label_question.config(text=question)
    nbSecondes=dict_question[question][2]    
    event=Event()
    thread1=Thread(target=minuteur,args=(event,nbSecondes,1))
    thread1.start()
    
    for l in labels:
        l.destroy()
    decalage=30
    for i in range(helpCompteur):
        label_coeur = tk.Label(fenetre, image=coeur)
        labels.append(label_coeur)      
        label_coeur.place(x=680+(i*decalage),y=0)
               
    afficher_suggestion(dict_question[question])

def afficher_suggestion(liste_suggestion):
    global intvar_reponses, bonne_reponses, indice_bonne_reponses

    clear_checkbuttons()
    intvar_reponses = []  
    bonne_reponses = liste_suggestion[1]
    indice_bonne_reponses=[]
    
    i=0
    random.shuffle(liste_suggestion[0])
    for elem in liste_suggestion[0]:
        if elem in liste_suggestion[1]:
            indice_bonne_reponses.append(i)
        var = tk.IntVar(value=0)  
        intvar_reponses.append(var)  
        checkB = tk.Checkbutton(checkbuttons_frame, text=elem, width=100,background='lightblue', variable=var, onvalue=1, offvalue=0, font=4,anchor="w", padx=220)

        checkB.pack()
        i=i+1

def suivant():
    global helpCompteur, index, nbQuestion, score, dict_reponse_fausse,note
    
    cpt=0
    cpt_b=0
    user_reponses=[]
    
    if flag==0:
        for i in range(len(intvar_reponses)):
            user_reponses.append(intvar_reponses[i].get())
  
        cpt=user_reponses.count(1)
        if cpt==len(bonne_reponses):
            for elem in indice_bonne_reponses:
                if user_reponses[elem] == 1: 
                    cpt_b=cpt_b+1
            if len(indice_bonne_reponses)==cpt_b:
                score=score+1
            else:
                dict_reponse_fausse[liste_question[index]]=user_reponses
        else:
            dict_reponse_fausse[liste_question[index]]=user_reponses
    else:
        nbQuestion+=1
        helpCompteur-=1
        if helpCompteur==0:
            label_changer.destroy()

    index += 1
    event.set()
                
    
    if index < nbQuestion:
        afficher_question(dict_question,liste_question, index)
    else:
        note=f"{score}/{total}"
        label_question.config(text=f"SCORE : {score}/{total}", foreground="green")
        clear_checkbuttons()
        label_bouton.destroy()
        label_compteur.destroy()
        label_changer.destroy()
        if score<index:
            correction=tk.Button(fenetre, text="Voir correction", command=voir_correction,width=50, cursor="hand2", bg="#98FB98", fg="black", pady=5)
            correction.pack()

def clear_checkbuttons():
    for widget in checkbuttons_frame.winfo_children():
        widget.destroy()
        
def structure(difficulter):
    global helpCompteur, label_changer, checkbuttons_frame, total, label_question, index,nbQuestion, label_bouton, liste_question, label_compteur, dict_question, dict_reponse_fausse,coeur,labels
       
    wb=load_workbook("QCM.xlsx")
    ws=wb.active
    
    index = 0
    nbQuestion=10
    total=10
    helpCompteur=4-difficulter
    labels=[]
    
    dict_question={}
        
    i=5    
    while ws[f"C{i}"].value!=None:
        cle=ws[f"C{i}"].value
        if ws[f"F{i}"].value==difficulter:
            dict_question[cle]=[[],[],0]
            dict_question[cle][2]=ws[f"E{i}"].value
            
            max=random.randint(2, 4)
            liste_indice=[i,i+1,i+2,i+3]
            random.shuffle(liste_indice)            
            n=0
            for indice in liste_indice:
                if n==max:
                    break                
                rep=ws[f"D{indice}"]
                if rep.value != None:
                    n=n+1
                    if rep.font.bold:
                        dict_question[cle][1].append(rep.value)
                    dict_question[cle][0].append(rep.value)    
            
        i=i+4        
    
    dict_reponse_fausse={}
    forme=font.Font(weight="bold")
    
    img = Image.open("coeur.png")
    img_resized = img.resize((25, 25))
    coeur = ImageTk.PhotoImage(img_resized)    
    
    label_compteur=tk.Label(fenetre, text="", foreground="red", background='lightblue', font=forme)
    label_compteur.pack()
   
    label_question = tk.Label(fenetre, text="", width=100, height=15, background='lightblue', font=forme,justify="left")
    label_question.pack()
    
    checkbuttons_frame = tk.Frame(fenetre, background='lightblue')
    checkbuttons_frame.pack()    

    label_bouton = tk.Button(fenetre, text="Suivant", width=60, command=suivant, cursor="hand2", bg="#98FB98", fg="black",pady=5)
    label_bouton.pack(pady=15)
    
    label_changer=tk.Button(fenetre, text="Changer Question", width=60, command=changer_question, cursor="hand2", bg="#98FB98", fg="black",pady=5)
    label_changer.pack()    

    liste_question=list(dict_question.keys())
    random.shuffle(liste_question)
    afficher_question(dict_question, liste_question, index)  

def changer_question():
    global flag
    flag=1
    suivant()
        
def voir_correction():
    for widget in fenetre.winfo_children():
        widget.destroy()    
        
    if num_difficulte==1:
        diff="Facile"
    elif num_difficulte==2:
        diff="Moyen"
    else:
        diff="Difficile"
    
    wb=Workbook()
    
    ws=wb.active
    ws.title='Correction'
       
    cellule(ws,"C1","Type",center,thick,bleu,True)    
    cellule(ws,"D1","QCM",center,thick,white,False) 
    cellule(ws,"C2","Difficulté",center,thick,bleu,True)
    cellule(ws,"D2",diff,center,thick,white,False)       
    cellule(ws,"C3","Score",center,thick,bleu,True)       
    cellule(ws,"D3",note,center,thick,white,False)
    cellule(ws,"C4","Question",center,thick,yellow,True)    
    cellule(ws,"D4","Votre réponse",center,thick,yellow,True)     
    cellule(ws,"E4","Réponse correcte",center,thick,yellow,True)
    
    ws.column_dimensions['C'].width=36
    ws.column_dimensions['D'].width=25
    ws.column_dimensions['E'].width=25
    ws.row_dimensions[4].height=30
           
    row=5
    for question,valeur in dict_reponse_fausse.items():
            ws[f"C{row}"]=question
            ws[f"C{row}"].alignment=left
            ws[f"C{row}"].border=thick
            if len(dict_question[question][1])==0:
                ws[f"E{row}"]="(Pas de réponse)"
            else: 
                ligne=[]
                for i in dict_question[question][1]:
                    ligne.append(str(i))
                ws[f"E{row}"]=" \n ".join(ligne)
            ws[f"E{row}"].alignment=center
            ws[f"E{row}"].border=thick
            ws[f"E{row}"].font=green
            j=0
            reponse_fausse=[]
            for indice in valeur:
                if indice==1:
                    reponse_fausse.append(str(dict_question[question][0][j]))
                j+=1        
            
            if reponse_fausse:    
                ws[f"D{row}"]=" \n ".join(reponse_fausse)
            else:
                ws[f"D{row}"]="(Pas de réponse)"
            ws[f"D{row}"].border=thick
            ws[f"D{row}"].alignment=center
            ws[f"D{row}"].font=red
            
            ws.row_dimensions[row].height=110  
            row+=1
    
    now = datetime.now()
    filename=name+' '+now.strftime("%Y-%m-%d %H-%M-%S")+'.xlsx'    
    wb.save(filename)
    os.system(f'start excel "{filename}"')

def passer(var):
        global num_difficulte
        
        num_difficulte=var
        label_difficulte.destroy()
        option1.destroy()
        option2.destroy()
        option3.destroy()
        structure(var)

def qcm():
    global label_difficulte,option1,option2,option3
    
    label_difficulte=tk.Label(text="Choisissez votre difficulté :", height=15, font=10, background='lightblue')
    label_difficulte.pack()
    
    option1 = tk.Button(fenetre, text="Facile", height=2, width=90, command=lambda:passer(1), cursor="hand2", bg="#98FB98", fg="black")
    option2 = tk.Button(fenetre, text="Moyen", height=2, width=75, command=lambda:passer(2), cursor="hand2", bg="#98FB98", fg="black")
    option3 = tk.Button(fenetre, text="Difficile", height=2, width=60, command=lambda:passer(3), cursor="hand2", bg="#98FB98", fg="black")
    option1.pack(pady=15)
    option2.pack(pady=15)
    option3.pack(pady=15)

def suivant_direct():
    global score,indice,event,dict_question,liste_question, liste_question, label_compteur,entry,user_reponse,signature,button,note_direct
    
    entry.destroy()
    
    user_reponse=user_reponse.get()
    reponse_correcte=eval(signature)
    
    if user_reponse==str(reponse_correcte):
        score=score+1
    else:
        dict_reponse_fausse[question]=(user_reponse,str(reponse_correcte))
    indice += 1
    
    event.set()
    
    if indice < len(list(dict_question.keys())):
        afficher_question_direct(dict_question,liste_question, indice)
    else:       
        button.destroy()
        label_compteur.destroy()
        note_direct=f'{score}/{indice}'
        label_question.config(text=f"SCORE : {score}/{indice}", foreground="green")
        if score<indice:
            correction=tk.Button(fenetre, text="Voir correction", command=voir_correction_direct,width=50, cursor="hand2", bg="#98FB98", fg="black", pady=5)
            correction.pack()
            
def modifier_element(ligne,i,n):
        ligne=list(ligne)
        ligne[i]=str(n)
        return "".join(ligne)   

def afficher_question_direct(dict_question, liste_question, index):
    global event,button,user_reponse,signature,question,entry
    question=liste_question[index]
    
    nomFonction=dict_question[question][0]

    liste_parametres=dict_question[question][1]
    i=0
    while i<len(liste_parametres):
        liste_parametres[i]=eval(liste_parametres[i])
        i=i+1
        
    cle=question
    j=0
    i=0
    n=len(question)
    while i<n:
        if question[i]=='$':
            m=liste_parametres[j]
            question=modifier_element(question,i,m)
            j=j+1
        i+=1 
        
    label_question.config(text=question)
    nbSecondes=dict_question[cle][2]
    
    event=Event()
    thread1=Thread(target=minuteur,args=(event,nbSecondes,2))
    thread1.start()    

    user_reponse=tk.StringVar()
    entry = tk.Entry(fenetre, width=20, textvariable=user_reponse, font=('Arial', 12), relief=tk.GROOVE, borderwidth=2, justify=tk.CENTER)
    entry.place(x=300,y=400)
    
    parametres_str=[]
    for par in liste_parametres:
        if type(par)==str:
            parametres_str.append(f"'{par}'")
        else:
            parametres_str.append(str(par))
    parametres_str=",".join(parametres_str)
    
    signature="fonctions."+nomFonction+"("+parametres_str+")"
 
def cellule(ws,reference,contenu,alignement,bordure,remplissage,gras):
    cell=ws[reference]
    cell.value=contenu
    cell.alignment=alignement
    cell.border=bordure
    cell.fill=remplissage    
    if gras:
        cell.font=bold
 
def voir_correction_direct():
    for widget in fenetre.winfo_children():
        widget.destroy()    
            
    wb=Workbook()
    
    ws=wb.active
    ws.title='Correction'
    
    cellule(ws,"C1","Type",center,thick,bleu,True)    
    cellule(ws,"D1","AVEC TROUS",center,thick,white,False)        
    cellule(ws,"C2","Score",center,thick,bleu,True)    
    cellule(ws,"D2",note_direct,center,thick,white,False)    
    cellule(ws,"C3","Question",center,thick,yellow,True)    
    cellule(ws,"D3","Votre réponse",center,thick,yellow,True)     
    cellule(ws,"E3","Réponse correcte",center,thick,yellow,True)
    
    ws.column_dimensions['C'].width=36
    ws.column_dimensions['D'].width=25
    ws.column_dimensions['E'].width=25
    ws.row_dimensions[3].height=30  
        
    row=4
    for question,valeur in dict_reponse_fausse.items():
            cellule(ws,f"C{row}",question,left,thick,white,False)
            cellule(ws,f"E{row}",valeur[1],center,thick,white,False)
            ws[f"E{row}"].font=green                                        
            
            if valeur[0]!="":    
                ws[f"D{row}"]=valeur[0]
            else:
                ws[f"D{row}"]="(Pas de réponse)"
            ws[f"D{row}"].border=thick
            ws[f"D{row}"].alignment=center
            ws[f"D{row}"].font=red
                                   
            ws.row_dimensions[row].height=110  
            row+=1
    
    now = datetime.now()
    filename=name+' '+now.strftime("%Y-%m-%d %H-%M-%S")+'.xlsx'    
    wb.save(filename)
    os.system(f'start excel "{filename}"')
    
def direct():
    global label_question,label_compteur,indice,dict_question,liste_question,dict_reponse_fausse,button
    wb=load_workbook("Direct.xlsx")
    ws=wb.active
    
    dict_question={}
    
    dict_reponse_fausse={}
    
    i=4
    while ws[f"C{i}"].value!=None:
        cle=ws[f"C{i}"].value
        nomFonction=ws[f"D{i}"].value
        ms=ws[f"F{i}"].value
        args=[]
        for j in range(i,i+4):
            valeur=ws[f"E{j}"].value
            if valeur != None:
                args.append(valeur)                        
        i=i+4      
        dict_question[cle]=(nomFonction,args,ms)
    
    forme=font.Font(weight="bold")
    
    label_compteur=tk.Label(fenetre, text="", foreground="red", background='lightblue', justify="left",font=forme)
    label_compteur.pack()

    label_question = tk.Label(fenetre, text="", width=100, height=15, background='lightblue', justify="left",font=forme)
    label_question.pack() 
    
    button=tk.Button(fenetre,text="Suivant",width=26,command=suivant_direct,cursor="hand2", bg="#98FB98", fg="black")
    button.place(x=300,y=480)   

    indice=0
    
    liste_question=list(dict_question.keys())
    random.shuffle(liste_question)
    
    afficher_question_direct(dict_question, liste_question, indice)
    
def choisir_type(num):
    label_type.destroy()
    type1.destroy()
    type2.destroy()
    if num==1:
        qcm()
    else:        
        direct()
        
score=0
name=input("Donner votre prénom : ")

fenetre = tk.Tk()
fenetre.title("QCM Python")
fenetre['bg'] = 'lightblue'
fenetre.geometry("800x1000")
fenetre.resizable(width=False, height=False)

label_type=tk.Label(text="Choississez le type de question :",height=15,font=10,background='lightblue')
label_type.pack()

type1=tk.Button(fenetre,text="QCM",height=2,width=80,command=lambda:choisir_type(1),cursor="hand2", bg="#98FB98", fg="black")
type2=tk.Button(fenetre,text="AVEC TROUS",height=2,width=80,command=lambda:choisir_type(2),cursor="hand2", bg="#98FB98", fg="black")
type1.pack(pady=15)
type2.pack()

fenetre.mainloop()